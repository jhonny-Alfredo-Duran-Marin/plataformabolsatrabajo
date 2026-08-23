from datetime import datetime
from io import BytesIO

from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.bitacora import BitacoraLog
from app.repositories.bitacora_repository import BitacoraRepository

_COLUMNAS = ("Fecha", "Usuario ID", "IP", "Módulo", "Acción", "Resultado", "Detalles")


class BitacoraService:
    def __init__(self, db: Session) -> None:
        self.repo = BitacoraRepository(db)

    def registrar(
        self,
        modulo: str,
        accion: str,
        usuario_id: int | None = None,
        ip: str | None = None,
        detalles: str | None = None,
        resultado: bool = True,
    ) -> BitacoraLog:
        log = BitacoraLog(
            usuario_id=usuario_id, ip=ip, modulo=modulo, accion=accion, detalles=detalles, resultado=resultado
        )
        return self.repo.registrar(log)

    def listar(
        self,
        usuario_id: int | None = None,
        modulo: str | None = None,
        accion: str | None = None,
        fecha_desde: datetime | None = None,
        fecha_hasta: datetime | None = None,
    ) -> list[BitacoraLog]:
        return self.repo.listar(
            usuario_id=usuario_id,
            modulo=modulo,
            accion=accion,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
        )

    @staticmethod
    def _fila(log: BitacoraLog) -> tuple:
        return (
            log.fecha.strftime("%Y-%m-%d %H:%M:%S") if log.fecha else "",
            log.usuario_id if log.usuario_id is not None else "",
            log.ip or "",
            log.modulo,
            log.accion,
            "Éxito" if log.resultado else "Fallo",
            log.detalles or "",
        )

    def exportar_excel(self, logs: list[BitacoraLog]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bitácora"
        ws.append(_COLUMNAS)
        for log in logs:
            ws.append(self._fila(log))

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def exportar_pdf(self, logs: list[BitacoraLog]) -> bytes:
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Bitácora de operaciones críticas - EGRESA", ln=True)

        pdf.set_font("Helvetica", "B", 9)
        anchos = (30, 20, 25, 30, 35, 20, 100)
        for columna, ancho in zip(_COLUMNAS, anchos):
            pdf.cell(ancho, 8, columna, border=1)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        for log in logs:
            for valor, ancho in zip(self._fila(log), anchos):
                texto = str(valor)[:80]
                pdf.cell(ancho, 7, texto, border=1)
            pdf.ln()

        return bytes(pdf.output())
