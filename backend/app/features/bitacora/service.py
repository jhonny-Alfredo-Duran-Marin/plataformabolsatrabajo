import uuid
from datetime import datetime
from io import BytesIO
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.seguridad import AuditLog
from app.features.bitacora.repository import BitacoraRepository

_COLUMNAS = ("Fecha", "Usuario ID", "IP", "Módulo", "Acción", "Resultado", "Detalles")


class BitacoraService:
    def __init__(self, db: Session) -> None:
        self.repo = BitacoraRepository(db)

    def registrar(
        self,
        modulo: str,
        accion: str,
        usuario_id: uuid.UUID | str | None = None,
        ip: str | None = None,
        detalles: str | None = None,
        resultado: bool = True,
    ) -> AuditLog:
        details: dict[str, Any] | None = {"detalle": detalles} if detalles else None
        log = AuditLog(
            user_id=usuario_id,
            ip_address=ip,
            entity_type=modulo,
            action=accion,
            result="success" if resultado else "failure",
            details_json=details,
        )
        return self.repo.registrar(log)

    def listar(
        self,
        usuario_id: uuid.UUID | str | None = None,
        modulo: str | None = None,
        accion: str | None = None,
        fecha_desde: datetime | None = None,
        fecha_hasta: datetime | None = None,
    ) -> list[AuditLog]:
        return self.repo.listar(
            usuario_id=usuario_id,
            modulo=modulo,
            accion=accion,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
        )

    @staticmethod
    def _fila(log: AuditLog) -> tuple:
        return (
            log.fecha.strftime("%Y-%m-%d %H:%M:%S") if log.fecha else "",
            log.user_id if log.user_id is not None else "",
            log.ip_address or "",
            log.entity_type,
            log.action,
            "Éxito" if log.resultado else "Fallo",
            log.detalles or "",
        )

    def exportar_excel(self, logs: list[AuditLog]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bitácora"
        ws.append(_COLUMNAS)
        for log in logs:
            ws.append(self._fila(log))

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def exportar_pdf(self, logs: list[AuditLog]) -> bytes:
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, "Bitácora de operaciones críticas - EGRESA", ln=True)

        pdf.set_font("Helvetica", "B", 9)
        anchos = (30, 20, 25, 30, 35, 20, 100)
        for columna, ancho in zip(_COLUMNAS, anchos):
            pdf.cell(ancho, 8, columna, border=1)
        pdf.ln()

        pdf.set_font("Helvetica", "", 8)
        for log in logs:
            for valor, ancho in zip(self._fila(log), anchos):
                texto = str(valor)[:80]
                pdf.cell(ancho, 7, texto, border=1)
            pdf.ln()

        return bytes(pdf.output())
